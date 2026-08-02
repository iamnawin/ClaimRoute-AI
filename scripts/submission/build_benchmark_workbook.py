"""Generate submission/final/05_Benchmark.xlsx from committed frozen evidence.

Every value written here comes from ``scripts/submission/evidence.py``, which reads only
committed artifacts. Nothing in this file hard-codes a benchmark number.

Run:  python scripts/submission/build_benchmark_workbook.py
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

from evidence import EVIDENCE, NOT_METERED, NOT_REPORTED  # noqa: E402

OUT = Path(__file__).resolve().parents[2] / "submission" / "final" / "05_Benchmark.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
NOTE_FONT = Font(italic=True, size=9, color="595959")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PCT = "0.000%"
USD6 = "$0.000000"
USD2 = "$#,##0.00"
MS = "#,##0.000"


def _title(ws, text: str, note: str | None = None) -> int:
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    row = 2
    if note:
        ws[f"A{row}"] = note
        ws[f"A{row}"].font = NOTE_FONT
        row += 1
    return row + 1


def _header(ws, row: int, headers: list[str]) -> int:
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    return row + 1


def _write(ws, row: int, values: list, fmts: dict[int, str] | None = None) -> int:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=isinstance(value, str))
        if fmts and col in fmts and isinstance(value, (int, float)):
            cell.number_format = fmts[col]
    return row + 1


def _widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _bar(ws, anchor: str, title: str, y_axis: str, data_ref, cats_ref, fmt: str | None = None):
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.y_axis.title = y_axis
    chart.height = 8
    chart.width = 17
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    if fmt:
        chart.y_axis.numFmt = fmt
    ws.add_chart(chart, anchor)


# --- sheets -----------------------------------------------------------------


def sheet_executive(wb: Workbook) -> None:
    ws = wb.create_sheet("Executive Metrics")
    row = _title(
        ws,
        "ClaimRoute AI - Executive Metrics",
        "Frozen synthetic test split. Every value is generated from eval/frozen/ evidence; "
        "synthetic and official results are never blended.",
    )
    row = _header(ws, row, ["Metric", "Value", "Evidence classification"])
    for metric, value, label in EVIDENCE.executive_metrics():
        fmt = None
        if isinstance(value, float):
            if "accuracy" in metric.lower() or metric in {"Precision", "Recall"} or "rate" in metric.lower():
                fmt = PCT
            elif "USD" in metric:
                fmt = USD6
            else:
                fmt = MS
        row = _write(ws, row, [metric, value, label], {2: fmt} if fmt else None)
    _widths(ws, [44, 30, 34])


def sheet_synthetic(wb: Workbook) -> None:
    ws = wb.create_sheet("Synthetic Benchmark")
    row = _title(
        ws,
        "Synthetic Benchmark - per degradation tier",
        "Clean, noisy, ugly and blended are reported separately with their own denominators. "
        "Escalation resolution uses the offline-oracle test double, not a real provider.",
    )
    headers = [
        "Scope", "Documents", "Pages", "Evaluated fields", "Routed fields",
        "Field accuracy", "Critical accuracy", "Primary local resolution",
        "Local retry rate", "Retry resolution", "Escalation rate",
        "Offline-oracle resolution", "Human review", "Accept with flag",
        "Measured local $/page", "Measured API $/page", "Projected API $/page",
        "Projected automated $/page",
    ]
    row = _header(ws, row, headers)
    fmts = {i: PCT for i in range(6, 15)}
    fmts.update({15: USD6, 16: USD6, 17: USD6, 18: USD6})

    chart_start = row
    for name in ("clean", "noisy", "ugly"):
        tier = EVIDENCE.per_tier[name]
        row = _write(ws, row, [
            name, tier["documents"], tier["pages"], tier["evaluated_fields"],
            tier["routed_fields"], tier["field_accuracy"], tier["critical_field_accuracy"],
            tier["primary_local_resolution_rate"], tier["local_retry_rate"],
            tier["local_retry_resolution_rate"], tier["escalation_rate"],
            tier["escalation_resolution_rate"], tier["human_review_rate"],
            tier["accept_with_flag_rate"], tier["measured_local_cost_per_page_usd"],
            tier["measured_api_spend_per_page_usd"], tier["projected_api_cost_per_page_usd"],
            tier["projected_total_automated_cost_per_page_usd"],
        ], fmts)
    b = EVIDENCE.blended
    row = _write(ws, row, [
        "blended", b["documents"], b["pages"], b["evaluated_fields"], b["routed_fields"],
        b["field_accuracy"], b["critical_field_accuracy"], b["primary_local_resolution_rate"],
        b["local_retry_rate"], b["local_retry_resolution_rate"], b["escalation_rate"],
        b["escalation_resolution_rate"], b["human_review_rate"], b["accept_with_flag_rate"],
        b["measured_local_cost_per_page_usd"], b["measured_api_spend_per_page_usd"],
        b["projected_api_cost_per_page_usd"], b["projected_total_automated_cost_per_page_usd"],
    ], fmts)

    _bar(
        ws, f"A{row + 2}", "Field accuracy by degradation tier", "Accuracy",
        Reference(ws, min_col=6, min_row=chart_start - 1, max_row=row - 1),
        Reference(ws, min_col=1, min_row=chart_start, max_row=row - 1),
        PCT,
    )
    _widths(ws, [12] + [13] * 17)


def sheet_official(wb: Workbook) -> None:
    ws = wb.create_sheet("Official Tier Evidence")
    row = _title(
        ws,
        "Official Organiser Dataset - Tier A / B / C / D reported separately",
        "Tier support differs, so no combined official accuracy is produced. Official results "
        "are never merged with the synthetic benchmark.",
    )
    row = _header(ws, row, [
        "Tier", "Document definition", "Support status", "Documents", "Pages",
        "Accuracy / measured result", "Evidence", "Holdout used", "External calls", "Limitations",
    ])
    for tier in EVIDENCE.official_tier_rows():
        row = _write(ws, row, [
            tier["tier"], tier["document"], tier["support"], tier["containers"], tier["pages"],
            tier["accuracy"], tier["evidence"], tier["holdout_used"], tier["external_calls"],
            tier["limitation"],
        ])
    _widths(ws, [7, 40, 21, 11, 8, 34, 44, 30, 13, 52])


def sheet_accuracy(wb: Workbook) -> None:
    ws = wb.create_sheet("Accuracy Metrics")
    counts = EVIDENCE.precision_recall()
    row = _title(
        ws,
        "Accuracy, Precision and Recall",
        "TP/FP/FN are derived field-by-field from eval/frozen/final_benchmark_rows.jsonl using "
        "the frozen accuracy denominator. Synthetic scope only.",
    )
    row = _header(ws, row, ["Metric", "Value", "Definition", "Evidence classification"])
    b = EVIDENCE.blended
    rows = [
        ("Exact field accuracy", b["field_accuracy"], "Normalized exact match over evaluated fields", "MEASURED SYNTHETIC"),
        ("Critical-field accuracy", b["critical_field_accuracy"], "Exact match over fields marked critical in frozen field policy", "MEASURED SYNTHETIC"),
        ("Automated exact-match rate", b["automated_exact_match_rate"], "Correct without any human review", "MEASURED SYNTHETIC"),
    ]
    if counts.derivable:
        rows += [
            ("True positives (TP)", counts.true_positives, "Required populated field extracted and matched ground truth", "MEASURED SYNTHETIC (derived)"),
            ("False positives (FP)", counts.false_positives, "Incorrect value accepted, or a value produced for a blank field", "MEASURED SYNTHETIC (derived)"),
            ("False negatives (FN)", counts.false_negatives, "Required populated field missed or left unresolved", "MEASURED SYNTHETIC (derived)"),
            ("Precision", counts.precision, "TP / (TP + FP)", "MEASURED SYNTHETIC (derived)"),
            ("Recall", counts.recall, "TP / (TP + FN)", "MEASURED SYNTHETIC (derived)"),
            ("Populated fields scored", counts.populated_fields, "Ground truth non-blank and in the frozen denominator", "MEASURED SYNTHETIC"),
            ("Blank-ground-truth fields scored", counts.blank_ground_truth_fields, "Legitimately empty; hallucinated values would count as FP", "MEASURED SYNTHETIC"),
        ]
    else:
        rows += [("Precision", NOT_REPORTED, "TP / (TP + FP)", "UNAVAILABLE"),
                 ("Recall", NOT_REPORTED, "TP / (TP + FN)", "UNAVAILABLE")]
    rows.append(("Human-review rate", b["human_review_rate"], "Fields routed to a human reviewer", "MEASURED SYNTHETIC"))

    for metric, value, definition, label in rows:
        fmt = PCT if isinstance(value, float) else None
        row = _write(ws, row, [metric, value, definition, label], {2: fmt} if fmt else None)
    _widths(ws, [34, 18, 62, 32])


def sheet_funnel(wb: Workbook) -> None:
    ws = wb.create_sheet("Resolution Funnel")
    row = _title(
        ws,
        "Resolution Funnel - where each field is actually resolved",
        "Applicable fields -> primary OCR -> local retry -> multimodal escalation -> human review. "
        "Shares are of routed fields.",
    )
    row = _header(ws, row, ["Stage", "Fields", "Share of routed fields", "Evidence classification"])
    chart_start = row
    for stage in EVIDENCE.resolution_funnel():
        row = _write(ws, row, [stage["stage"], stage["fields"], stage["share_of_routed"], stage["label"]], {3: PCT})
    _bar(
        ws, f"A{row + 2}", "Resolution funnel (fields by stage)", "Fields",
        Reference(ws, min_col=2, min_row=chart_start - 1, max_row=row - 1),
        Reference(ws, min_col=1, min_row=chart_start, max_row=row - 1),
    )
    _widths(ws, [40, 12, 22, 42])


def sheet_cost(wb: Workbook) -> None:
    ws = wb.create_sheet("Cost Breakdown")
    row = _title(
        ws,
        "Cost Breakdown per page",
        "Local stages are measured usage priced at a configured compute rate: near-zero "
        "incremental cost, never free. External API spend is measured at $0 because no paid "
        "call was made; the selective AI figure is projected from offline-oracle token counts.",
    )
    row = _header(ws, row, ["Component", "Calls in run", "Total USD", "USD per page", "Evidence classification"])

    labels = {
        "preprocess": ("Tier-0 preprocessing", "MEASURED local compute"),
        "route": ("Document routing", "MEASURED local compute"),
        "ocr_paddle": ("Primary OCR (PaddleOCR/RapidOCR)", "MEASURED local compute"),
        "validate_fuse": ("Validation and confidence fusion", "MEASURED local compute"),
        "retry_tesseract": ("Local crop retry OCR (Tesseract)", "MEASURED local compute"),
        "escalate_intent": ("Escalation decision (governor)", "MEASURED local compute"),
        "escalate_offline-oracle": ("Selective multimodal escalation", "PROJECTED OFFLINE_ORACLE"),
    }
    chart_start = row
    for stage in EVIDENCE.stage_costs():
        name, label = labels.get(stage["operation"], (stage["operation"], "MEASURED"))
        row = _write(ws, row, [name, stage["calls"], stage["cost_usd"], stage["cost_per_page_usd"], label],
                     {3: USD6, 4: USD6})
    chart_end = row - 1

    b = EVIDENCE.blended
    row += 1
    for name, value, label in [
        ("CPU compute allocation", b["measured_local_cost_per_page_usd"], "MEASURED usage at ASSUMED $/vCPU-hour"),
        ("GPU compute", NOT_METERED, "Pipeline ran CPU-only; no GPU stage exists"),
        ("LLM (text-only) cost", NOT_METERED, "No text-only LLM stage exists in the pipeline"),
        ("Vision AI cost", b["projected_api_cost_per_page_usd"], "PROJECTED OFFLINE_ORACLE"),
        ("Measured external API invoice", b["measured_external_api_spend_usd"], "MEASURED; zero paid calls"),
        ("Total automated cost per page", b["projected_total_automated_cost_per_page_usd"], "PROJECTED local + selective AI"),
        ("Human review", 0.03, "ASSUMED configurable illustrative rate per reviewed field; reported separately"),
    ]:
        fmt = {4: USD6} if isinstance(value, (int, float)) else None
        row = _write(ws, row, [name, "", "", value, label], fmt)

    row += 1
    row = _write(ws, row, ["Volume projection", "Local compute", "Selective AI", "Automated total", "Human review assumption"])
    for proj in EVIDENCE.cost_projection():
        row = _write(ws, row, [
            f"{int(proj['pages']):,} pages",
            float(proj["measured_local_compute_projection_usd"]),
            float(proj["projected_api_cost_usd"]),
            float(proj["projected_total_automated_cost_usd"]),
            float(proj["configured_human_review_cost_usd"]),
        ], {2: USD2, 3: USD2, 4: USD2, 5: USD2})

    _bar(
        ws, f"G{chart_start}", "Cost per page by pipeline stage", "USD per page",
        Reference(ws, min_col=4, min_row=chart_start - 1, max_row=chart_end),
        Reference(ws, min_col=1, min_row=chart_start, max_row=chart_end),
        USD6,
    )
    _widths(ws, [38, 14, 14, 16, 50])


def sheet_latency(wb: Workbook) -> None:
    ws = wb.create_sheet("Latency and Throughput")
    t = EVIDENCE.throughput
    b = EVIDENCE.blended
    env = t.get("environment_manifest", "eval/frozen/environment_manifest.json")
    row = _title(
        ws,
        "Latency and Throughput",
        "Single-worker development-workstation measurement. This is not a production SLA; "
        "production scaling is an architectural projection.",
    )
    row = _header(ws, row, ["Metric", "Value", "Unit", "Evidence classification"])
    for metric, value, unit, label in [
        ("Pages measured", t["pages"], "pages", "MEASURED"),
        ("Total processing time", round(t["total_elapsed_ms"] / 1000.0, 3), "seconds", "MEASURED"),
        ("Average latency", t["average_latency_per_page_ms"], "ms/page", "MEASURED"),
        ("P50 latency", t["p50_latency_ms"], "ms", "MEASURED"),
        ("P95 latency", t["p95_latency_ms"], "ms", "MEASURED"),
        ("Pages per second", round(t["pages_per_minute"] / 60.0, 6), "pages/s", "MEASURED"),
        ("Pages per minute", t["pages_per_minute"], "pages/min", "MEASURED"),
        ("Pages per hour", t["pages_per_hour"], "pages/hr", "MEASURED"),
        ("Concurrency", 1, "worker", "MEASURED single-worker"),
        ("Environment manifest", env, "path", "MEASURED"),
        ("Warm-up handling", t["warm_up"], "policy", "MEASURED"),
        ("Memory", t["memory_observation"], "n/a", "NOT SEPARATELY METERED"),
        ("Provider latency", t["projected_provider_latency"], "n/a", "NOT AVAILABLE - offline oracle is not a provider"),
    ]:
        fmt = MS if isinstance(value, float) else None
        row = _write(ws, row, [metric, value, unit, label], {2: fmt} if fmt else None)

    row += 1
    row = _write(ws, row, ["Per-tier latency", "Average ms/page", "P50 ms", "P95 ms"])
    chart_start = row
    for name in ("clean", "noisy", "ugly"):
        tier = EVIDENCE.per_tier[name]
        row = _write(ws, row, [name, tier["average_latency_per_page_ms"], tier["p50_latency_ms"], tier["p95_latency_ms"]],
                     {2: MS, 3: MS, 4: MS})
    _bar(
        ws, f"F{chart_start}", "Average latency by degradation tier", "ms per page",
        Reference(ws, min_col=2, min_row=chart_start - 1, max_row=row - 1),
        Reference(ws, min_col=1, min_row=chart_start, max_row=row - 1),
    )
    ws.cell(row=row + 1, column=1, value=f"Blended average: {b['average_latency_per_page_ms']:.3f} ms/page").font = NOTE_FONT
    _widths(ws, [30, 34, 14, 46])


def sheet_frontier(wb: Workbook) -> None:
    ws = wb.create_sheet("Accuracy-Cost Frontier")
    points = EVIDENCE.frontier()
    row = _title(
        ws,
        "Accuracy-Cost Frontier (Economy / Balanced / Accuracy)",
        "SYNTHETIC replay-calibration evidence over recorded escalation candidates. It reran no "
        "OCR and called no provider, and it does not replace the frozen test result.",
    )
    if not points:
        _write(ws, row, ["No frozen calibration evidence found; frontier deliberately omitted."])
        _widths(ws, [80])
        return

    named = {
        "local-0.80_model-0.88_paid-high_flags-on": "Economy",
        "local-0.80_model-0.90_paid-high-med_flags-on": "Balanced",
        "local-0.88_model-0.90_paid-high-med-low_flags-on": "Accuracy",
    }
    row = _header(ws, row, [
        "Mode", "Calibration point", "Field accuracy", "Critical accuracy",
        "Escalation rate", "Human-review rate", "Projected automated $/page", "Evidence classification",
    ])
    chart_start = row
    for point in points:
        label = named.get(point["point_id"], "(sweep point)")
        row = _write(ws, row, [
            label, point["point_id"], float(point["field_accuracy"]),
            float(point["critical_field_accuracy"]), float(point["escalation_rate"]),
            float(point["human_review_rate"]), float(point["projected_total_automated_cost_per_page_usd"]),
            "SYNTHETIC replay calibration; PROJECTED cost",
        ], {3: PCT, 4: PCT, 5: PCT, 6: PCT, 7: USD6})
    _bar(
        ws, f"A{row + 2}", "Accuracy across calibration points", "Field accuracy",
        Reference(ws, min_col=3, min_row=chart_start - 1, max_row=row - 1),
        Reference(ws, min_col=1, min_row=chart_start, max_row=row - 1),
        PCT,
    )
    _widths(ws, [14, 48, 15, 16, 15, 18, 26, 40])


def sheet_coverage(wb: Workbook) -> None:
    ws = wb.create_sheet("A_B_C_D Coverage")
    row = _title(
        ws,
        "Tier A / B / C / D Coverage and Support Classification",
        "Support classification states what is genuinely proven per tier. FULLY_BENCHMARKED is "
        "claimed only where a frozen holdout result exists.",
    )
    row = _header(ws, row, ["Tier", "Document definition", "Support classification", "What is proven", "What is not proven"])
    chart_start = row
    for tier in EVIDENCE.official_tier_rows():
        row = _write(ws, row, [tier["tier"], tier["document"], tier["support"], tier["evidence"], tier["limitation"]])
    rank = {"FULLY_BENCHMARKED": 4, "PARTIAL_EVIDENCE": 3, "DEVELOPMENT_PROOF": 2, "ROUTING_ONLY": 1}
    ws.cell(row=chart_start - 1, column=6, value="Support level (4=highest)")
    for offset, tier in enumerate(EVIDENCE.official_tier_rows()):
        ws.cell(row=chart_start + offset, column=6, value=rank[tier["support"]]).border = BORDER
    _bar(
        ws, f"A{row + 2}", "Evidence strength by organiser tier", "Support level",
        Reference(ws, min_col=6, min_row=chart_start - 1, max_row=row - 1),
        Reference(ws, min_col=1, min_row=chart_start, max_row=row - 1),
    )
    _widths(ws, [7, 40, 22, 46, 52, 22])


def sheet_methodology(wb: Workbook) -> None:
    ws = wb.create_sheet("Methodology")
    integrity = EVIDENCE.integrity
    row = _title(ws, "Methodology", "How every number in this workbook was produced.")
    row = _header(ws, row, ["Topic", "Method"])
    for topic, method in [
        ("Dataset source", "Fully synthetic CMS-1500 / UB-04 generated by data_factory/ from seed 42. Zero PHI by construction."),
        ("Synthetic vs official separation", "The synthetic frozen benchmark and the official organiser dataset are reported in separate sheets and are never blended into one score."),
        ("Frozen split", f"Claim-level split pinned by sha256 {integrity['test_sha256']}. Calibration/test overlap = {integrity['calibration_test_overlap']}. Thresholds were never tuned on the test split."),
        ("Frozen commit", f"All evidence originates from git commit {EVIDENCE.frozen_commit}."),
        ("Normalization rules", "Field-aware typed normalization (dates, money, quantities, codes, identifiers) applied identically to candidate and ground truth before comparison."),
        ("Field denominators", f"{integrity['frozen_test_fields']} evaluated fields across {integrity['frozen_test_pages']} pages. Legitimately-empty optional fields are excluded when absent, and are counted as errors if a value is hallucinated into them."),
        ("Critical-field definition", "Fields marked high criticality in the frozen configs/field_policy.yaml."),
        ("Accuracy definition", "Normalized exact match. Partial credit is never awarded."),
        ("Precision / recall definition", "TP = required populated field extracted and matched. FP = incorrect value accepted, or value produced for a blank field. FN = required populated field missed or unresolved. Derived per field from the frozen rows."),
        ("Cost accounting", "Every stage writes to an append-only ledger. Local stages are measured usage priced at a configured compute rate. External spend is measured separately and was $0 because no paid call was made."),
        ("Throughput methodology", "Single worker, one process, development workstation. First page retained, no warm-up exclusion. Not a production SLA."),
        ("Offline-oracle limitation", "The offline oracle is a deterministic test double for the escalation boundary. Its accuracy is not evidence about any real model, and its cost is projected from token counts, never measured spend."),
        ("Human-review assumption", "Human review cost is a configurable illustrative assumption reported separately from automated cost, never folded into cost per page."),
        ("Leakage prevention", "Calibration and test splits are disjoint and verified; duplicate field rows and duplicate page rows both measured zero."),
        ("Holdout discipline", "The official Tier C holdout was authorized once and is consumed. It must not be rerun. Tier A holdout remains unopened."),
        ("Components not metered", f"GPU and text-only LLM costs are reported as {NOT_METERED} because no such stage exists in the pipeline. Memory was not measured."),
    ]:
        row = _write(ws, row, [topic, method])
    _widths(ws, [30, 128])


def sheet_assumptions(wb: Workbook) -> None:
    ws = wb.create_sheet("Assumptions and Limitations")
    row = _title(ws, "Assumptions and Limitations", "Stated plainly, because a judge will test these.")
    row = _header(ws, row, ["Item", "Honest boundary", "Consequence"])
    for item, boundary, consequence in [
        ("Synthetic benchmark", "The headline accuracy is synthetic only", "Real-claim generalization is unverified"),
        ("Official Tier C", "Provisional visible-and-supported denominator; three documents", "Not a universal UB-04 benchmark"),
        ("Official Tier A", "Development proof only; holdout unopened", "No official Tier A accuracy claim is made"),
        ("Official Tier D", "Conservative extraction; no confirmed denominator", "Tier D extraction capability is not claimed"),
        ("Real providers", "No paid provider smoke test was completed", "No provider accuracy or latency claim"),
        ("Offline oracle", "Deterministic test double, not a model", "Its resolution rate is not model performance"),
        ("External spend", "Measured at $0 because zero paid calls were made", "The selective AI figure is projected, not invoiced"),
        ("Compute price", "$/vCPU-hour is a configured assumption", "Local cost scales with the rate actually paid"),
        ("Human review", "$0.03 per reviewed field is illustrative and configurable", "Reported separately; never folded into cost per page"),
        ("GPU / LLM metering", NOT_METERED, "No GPU or text-only LLM stage exists to meter"),
        ("Throughput", "Single worker on a development workstation", "No production SLA or capacity guarantee"),
        ("Volume projections", "Linear extrapolation", "Excludes storage, orchestration, redundancy, observability, networking, support and tax"),
        ("Production architecture", "Designed, not deployed", "Enterprise components are a roadmap, not running infrastructure"),
        ("Security", "PHI-minimizing prototype controls", "No HIPAA certification or compliance claim"),
        ("UI scope", "Local workspace processes raster pages and PDFs; no production queue", "Batch orchestration at enterprise scale is roadmap"),
    ]:
        row = _write(ws, row, [item, boundary, consequence])
    _widths(ws, [26, 58, 62])


def sheet_evidence_index(wb: Workbook) -> None:
    ws = wb.create_sheet("Evidence Index")
    row = _title(ws, "Evidence Index", "Every number in this workbook traces to one of these committed files.")
    row = _header(ws, row, ["Artifact", "Contents", "Used by sheet"])
    for artifact, contents, used in [
        ("eval/frozen/frozen_manifest.json", "Frozen commit, split, policy, commands, evidence labels", "Executive Metrics, Methodology"),
        ("eval/frozen/final_benchmark_summary.json", "Blended and per-tier accuracy, routing, cost, latency", "Executive Metrics, Synthetic Benchmark, Accuracy Metrics"),
        ("eval/frozen/final_benchmark_rows.jsonl", "Field-level candidates, ground truth, decisions", "Accuracy Metrics (TP/FP/FN derivation)"),
        ("eval/frozen/final_benchmark_pages.jsonl", "Page receipts, counts, cost, latency", "Latency and Throughput"),
        ("eval/frozen/final_benchmark_ledger.jsonl", "Append-only per-operation cost ledger", "Cost Breakdown"),
        ("eval/frozen/throughput_summary.json", "Measured local prototype throughput", "Latency and Throughput"),
        ("eval/frozen/ablation_summary.json", "Supported ablation arms and unsupported-arm disclosure", "Methodology"),
        ("eval/frozen/cost_projection.csv", "1K to 100M page volume projections", "Cost Breakdown"),
        ("eval/frozen/config_hashes.json", "SHA-256 of configs and critical runtime files", "Methodology"),
        ("eval/frozen/environment_manifest.json", "Python, OCR, platform and package versions", "Latency and Throughput"),
        ("eval/results/day9_accuracy_cost_frontier.csv", "Replay calibration sweep", "Accuracy-Cost Frontier"),
        ("eval/results/official_ub04_holdout_summary.json", "One-time official Tier C receipt", "Official Tier Evidence, A_B_C_D Coverage"),
        ("eval/official/results/official_sample_summary.json", "PHI-safe official Tier A/B/C/D sample receipt", "Official Tier Evidence, A_B_C_D Coverage"),
        ("docs/submission/EVIDENCE_REGISTER.md", "Claim-to-evidence control register", "All sheets"),
    ]:
        row = _write(ws, row, [artifact, contents, used])
    _widths(ws, [48, 62, 52])


def main() -> int:
    wb = Workbook()
    wb.remove(wb.active)
    sheet_executive(wb)
    sheet_synthetic(wb)
    sheet_official(wb)
    sheet_accuracy(wb)
    sheet_funnel(wb)
    sheet_cost(wb)
    sheet_latency(wb)
    sheet_frontier(wb)
    sheet_coverage(wb)
    sheet_methodology(wb)
    sheet_assumptions(wb)
    sheet_evidence_index(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT} ({OUT.stat().st_size:,} bytes, {len(wb.sheetnames)} sheets)")
    for name in wb.sheetnames:
        print(f"  - {name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
