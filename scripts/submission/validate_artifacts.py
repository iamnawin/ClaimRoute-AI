"""Deep validation of the three generated submission artifacts.

Read-only. Opens each artifact the way a reviewer's tool would, then checks that
the numbers and labels the evidence register requires are actually present in the
rendered output rather than merely present in the generator source.

Uses only reportlab/openpyxl's own dependencies plus the standard library, so it
adds nothing to requirements.txt.

Exit code 0 when every check passes, 1 otherwise. Prints "<n>/<n> checks passed"
on success, which scripts/validate_submission_readiness.ps1 greps for.
"""
from __future__ import annotations

import base64
import re
import zipfile
import zlib
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
FINAL = ROOT / "submission" / "final"

EXEC_PDF = FINAL / "01_Executive_Summary.pdf"
ARCH_PDF = FINAL / "02_Architecture.pdf"
BENCH_XLSX = FINAL / "05_Benchmark.xlsx"

EXPECTED_SHEETS = [
    "Executive Metrics",
    "Synthetic Benchmark",
    "Official Tier Evidence",
    "Accuracy Metrics",
    "Resolution Funnel",
    "Cost Breakdown",
    "Latency and Throughput",
    "Accuracy-Cost Frontier",
    "A_B_C_D Coverage",
    "Methodology",
    "Assumptions and Limitations",
    "Evidence Index",
]

# The corrected derivation: TP=3106, FP=3, FN=59. 98.043% is the exact-match rate
# and must never appear in a cell or sentence labelled Recall.
PRECISION = 3106 / 3109
RECALL = 3106 / 3165
PRECISION_TEXT = "99.9"
RECALL_TEXT = "98.1"
STALE_RECALL = "98.043"

failures: list[str] = []
checks = 0


def check(label: str, ok: bool, detail: str = "") -> None:
    global checks
    checks += 1
    if ok:
        print(f"  PASS  {label}")
    else:
        print(f"  FAIL  {label}{(' - ' + detail) if detail else ''}")
        failures.append(label)


def _decode_stream(body: bytes) -> bytes | None:
    """Undo reportlab's stream filters.

    reportlab writes ``/Filter [ /ASCII85Decode /FlateDecode ]``, so the ASCII85
    layer has to come off before inflating. Plain Flate is also accepted so the
    check does not depend on which filter chain reportlab picks.
    """
    candidates = [body]
    if b"~>" in body:
        try:
            candidates.insert(0, base64.a85decode(body.strip(), adobe=True))
        except ValueError:
            pass
    for candidate in candidates:
        try:
            return zlib.decompressobj().decompress(candidate)
        except zlib.error:
            continue
    return None


def pdf_text(path: Path) -> str:
    """Extract visible text from the page content streams.

    Needs no PDF library. Text is returned with escapes collapsed; it is used for
    substring checks only, never for rendering.
    """
    raw = path.read_bytes()
    chunks: list[str] = []
    for match in re.finditer(rb"stream\r?\n(.*?)endstream", raw, re.S):
        decoded = _decode_stream(match.group(1))
        if decoded is None:
            continue
        for token in re.finditer(rb"\((?:[^()\\]|\\.)*\)", decoded):
            piece = token.group(0)[1:-1]
            piece = re.sub(rb"\\([()\\])", rb"\1", piece)
            chunks.append(piece.decode("latin-1"))
    return " ".join(chunks)


def validate_pdf(path: Path, label: str, must_contain: list[str]) -> None:
    if not path.exists():
        check(f"{label}: present", False, str(path))
        return
    check(f"{label}: present", True)

    raw = path.read_bytes()
    check(f"{label}: non-zero size", len(raw) > 0)
    check(f"{label}: PDF header", raw[:5] == b"%PDF-", raw[:8].decode("latin-1", "replace"))
    check(f"{label}: EOF marker", b"%%EOF" in raw[-2048:])

    pages = len(re.findall(rb"/Type\s*/Page[^s]", raw))
    check(f"{label}: has pages", pages > 0, f"found {pages}")

    text = pdf_text(path)
    check(f"{label}: text extraction succeeds", len(text) > 500, f"{len(text)} chars")

    for needle in must_contain:
        check(f"{label}: contains {needle!r}", needle in text)


def validate_workbook() -> None:
    if not BENCH_XLSX.exists():
        check("Benchmark: present", False, str(BENCH_XLSX))
        return
    check("Benchmark: present", True)

    check("Benchmark: valid zip container", zipfile.is_zipfile(BENCH_XLSX))
    with zipfile.ZipFile(BENCH_XLSX) as archive:
        bad = archive.testzip()
        check("Benchmark: no corrupt members (opens without repair)", bad is None, str(bad))
        names = archive.namelist()
        charts = [n for n in names if n.startswith("xl/charts/chart")]
        check("Benchmark: contains charts", len(charts) > 0, f"{len(charts)} chart parts")

    try:
        from openpyxl import load_workbook
    except ImportError:
        check("Benchmark: openpyxl available", False, "pip install openpyxl==3.1.5")
        return
    check("Benchmark: openpyxl available", True)

    workbook = load_workbook(BENCH_XLSX, data_only=False)
    sheets = workbook.sheetnames
    check("Benchmark: 12 sheets", len(sheets) == 12, f"found {len(sheets)}")
    for name in EXPECTED_SHEETS:
        check(f"Benchmark: sheet {name!r}", name in sheets)

    values: list[str] = []
    metrics: dict[str, list[float]] = {"Precision": [], "Recall": []}
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    values.append(str(cell))
            label = str(row[0]).strip() if row and row[0] is not None else ""
            if label in metrics:
                # Rates are stored as raw floats and shown through a percent
                # number format, so compare the value, not its rendering.
                metrics[label].extend(c for c in row[1:] if isinstance(c, float))
    blob = " ".join(values)

    check("Benchmark: reports Precision", "Precision" in blob)
    check("Benchmark: reports Recall", "Recall" in blob)
    check(
        "Benchmark: precision value is the derived 0.99903506",
        bool(metrics["Precision"]) and all(abs(v - PRECISION) < 5e-9 for v in metrics["Precision"]),
        str(metrics["Precision"]),
    )
    check(
        "Benchmark: recall value is the derived 0.98135861",
        bool(metrics["Recall"]) and all(abs(v - RECALL) < 5e-9 for v in metrics["Recall"]),
        str(metrics["Recall"]),
    )

    # Provenance labels must survive into the rendered workbook.
    lowered = blob.lower()
    for label in ("MEASURED", "PROJECTED", "SYNTHETIC", "OFFICIAL"):
        check(f"Benchmark: {label} label retained", label.lower() in lowered)

    # Synthetic and official evidence stay on separate sheets.
    check(
        "Benchmark: official evidence has its own sheet",
        "Official Tier Evidence" in sheets and "Synthetic Benchmark" in sheets,
    )


def validate_no_stale_recall() -> None:
    """98.043% is the exact-match rate. It must not be presented as recall."""
    for path, label in ((EXEC_PDF, "Executive Summary"), (ARCH_PDF, "Architecture")):
        if not path.exists():
            continue
        text = pdf_text(path)
        # Only the percentage bound to the Recall label itself. The exact-match
        # rate legitimately prints 98.043% on a neighbouring row.
        bound = re.findall(r"Recall\s+([0-9]+\.[0-9]+)%", text)
        offending = [v for v in bound if v.startswith(STALE_RECALL)]
        check(
            f"{label}: recall is not the stale {STALE_RECALL}%",
            not offending,
            f"Recall reads {offending[0]}%" if offending else "",
        )
        if bound:
            check(
                f"{label}: recall reads 98.136%",
                all(v.startswith("98.13") for v in bound),
                str(bound),
            )

    if BENCH_XLSX.exists():
        try:
            from openpyxl import load_workbook
        except ImportError:
            return
        workbook = load_workbook(BENCH_XLSX, data_only=False)
        offending = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if any("Recall" in c for c in cells) and any(STALE_RECALL in c for c in cells):
                    offending.append(cells)
        check(f"Benchmark: recall is not the stale {STALE_RECALL}%", not offending, str(offending[:1]))


def main() -> int:
    print("Deep artifact validation - submission/final")
    print()

    print("01_Executive_Summary.pdf")
    validate_pdf(EXEC_PDF, "Executive Summary", ["ClaimRoute", "Precision", "Recall", PRECISION_TEXT, RECALL_TEXT])
    print()

    print("02_Architecture.pdf")
    validate_pdf(ARCH_PDF, "Architecture", ["ClaimRoute", "Precision", "Recall"])
    print()

    print("05_Benchmark.xlsx")
    validate_workbook()
    print()

    print("Cross-artifact consistency")
    validate_no_stale_recall()
    print()

    if failures:
        print(f"{checks - len(failures)}/{checks} checks passed, {len(failures)} FAILED")
        for name in failures:
            print(f"  - {name}")
        return 1

    print(f"{checks}/{checks} checks passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
